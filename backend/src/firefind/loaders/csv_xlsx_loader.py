# loaders/csv_xlsx_loader.py
from typing import Dict, Iterator, List
from pathlib import Path
import csv
from openpyxl import load_workbook

from ..vendors.utils import normalize_header


# Heuristic markers that indicate a row is being used as the header line in
# firewall exports.  Some of the sample CSV files contain banner text and other
# metadata before the actual column headings which confused the old
# ``csv.DictReader`` logic because it would treat that banner row as the header
# and return dictionaries containing only the empty string key.  By recognising
# a few characteristic combinations of column names we can reliably identify the
# real header row and skip any preamble.
HEADER_MARKERS = [
    {"seq_#", "action", "service"},
    {"policyid", "srcaddr", "dstaddr"},
    {"id", "srcaddr", "dstaddr"},
]


def _read_csv_rows(path: Path) -> Iterator[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = list(csv.reader(f))

    if not reader:
        return

    header_idx = None
    fallback_idx = None

    for idx, row in enumerate(reader):
        cells = [str(cell or "").strip() for cell in row]
        if not any(cells):
            continue

        if fallback_idx is None:
            fallback_idx = idx

        normalised = {normalize_header(cell) for cell in cells if cell}
        if normalised and any(marker <= normalised for marker in HEADER_MARKERS):
            header_idx = idx
            break

    if header_idx is None:
        if fallback_idx is None:
            return
        header_idx = fallback_idx

    header = [str(cell or "").strip() for cell in reader[header_idx]]

    for row in reader[header_idx + 1 :]:
        values = [str(cell or "").strip() for cell in row]
        if len(values) < len(header):
            values += [""] * (len(header) - len(values))
        row_dict = {
            (header[i] or "").strip(): values[i] if values[i] is not None else ""
            for i in range(len(header))
        }

        # Skip section headings and noise only when a non-numeric Seq # is present
        seq = row_dict.get("Seq #")
        if seq is not None:
            seq = seq.strip()
            if seq and not seq.isdigit():
                continue

        yield row_dict


def _read_xlsx_rows(path: Path) -> Iterator[Dict[str, str]]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active

        # find the header row
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if {"Seq #", "Action", "Service"} <= set(cells):
                header_row_idx = i
                headers = cells
                break

        if header_row_idx is None:
            # fallback: take the first non-empty row as headers
            ws = wb.active
            ws.reset_dimensions()
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    header_row_idx = i
                    headers = cells
                    break

        if header_row_idx is None:
            return

        # Backfill empty header cells using the previous row when available. Many
        # vendor exports repeat banner metadata above the real table headings but
        # place labels such as "Risk Rating" there.  Without this step those
        # values would be lost because the actual header row contains blank
        # cells at the same positions.
        if header_row_idx > 1:
            previous = next(
                ws.iter_rows(
                    min_row=header_row_idx - 1,
                    max_row=header_row_idx - 1,
                    values_only=True,
                )
            )
            previous_cells = [
                str(cell).strip() if cell is not None else "" for cell in previous
            ]
            headers = [
                header if (header or "").strip() else previous_cells[idx]
                if idx < len(previous_cells)
                else ""
                for idx, header in enumerate(headers)
            ]

        # Propagate the last seen non-empty header across blank cells so the
        # loader produces names like ``Service.1`` that mirror pandas' naming
        # convention.  Several mappings rely on those derived labels when
        # looking up auxiliary columns such as port definitions.
        propagated_headers: List[str] = []
        previous_header = ""
        propagated_counts: Dict[str, int] = {}
        for header in headers:
            current = (header or "").strip()
            if current:
                previous_header = current
                propagated_headers.append(current)
                continue

            if not previous_header:
                propagated_headers.append("")
                continue

            counter = propagated_counts.get(previous_header, 0) + 1
            propagated_counts[previous_header] = counter
            propagated_headers.append(f"{previous_header}.{counter}")

        headers = propagated_headers

        # Ensure header names remain unique so trailing columns are not
        # collapsed when converted to a dictionary.
        seen_headers: Dict[str, int] = {}
        normalised_headers: List[str] = []
        for idx, header in enumerate(headers):
            base = (header or "").strip()
            counter = seen_headers.get(base, 0)
            seen_headers[base] = counter + 1
            if counter:
                normalised_headers.append(f"{base}__{counter}")
            else:
                normalised_headers.append(base)

        headers = normalised_headers

        # iterate rows after header
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            values = [("" if v is None else str(v).strip()) for v in row]
            # pad to headers length
            if len(values) < len(headers):
                values += [""] * (len(headers) - len(values))
            row_dict = dict(zip(headers, values))

            # Skip section headings and noise only when a non-numeric Seq # is present
            seq = row_dict.get("Seq #")
            if seq is not None:
                seq = seq.strip()
                if not seq.isdigit():
                    continue

            yield row_dict
    finally:
        wb.close()


def load_table(path: Path) -> Iterator[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _read_csv_rows(path)
    elif suffix == ".xlsx":
        yield from _read_xlsx_rows(path)
    else:
        raise ValueError(f"Unsupported file: {path}")
