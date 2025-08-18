import csv
from typing import Dict, Iterable, List
from openpyxl import load_workbook

def load_table(path: str) -> Iterable[Dict[str, str]]:
    """Yield rows as dicts from CSV or XLSX based on file extension."""
    path_lower = path.lower()
    if path_lower.endswith(".csv"):
        with open(path, newline='', encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield { (k or '').strip(): (v or '').strip() for k, v in row.items() }
    elif path_lower.endswith(".xlsx"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() for h in next(rows)]
        for r in rows:
            row = { headers[i]: (str(r[i]).strip() if r[i] is not None else "") for i in range(len(headers)) }
            yield row
    else:
        raise ValueError(f"Unsupported file type for {path}. Use .csv or .xlsx")
